# # # # from fastapi import FastAPI, HTTPException
# # # # from pydantic import BaseModel
# # # # from dotenv import load_dotenv
# # # # import os

# # # # # Import modules from our app directory
# # # # from app.email_parser import parse_email
# # # # from app.web_scraper import get_full_url, scrape_website_content
# # # # from app.llm_processor import LLMProcessor
# # # # from app.email_generator import EmailGenerator
# # # # from app.email_sender import EmailSender

# # # # # Load environment variables from .env file
# # # # load_dotenv()

# # # # app = FastAPI(
# # # #     title="Email Automation API",
# # # #     description="An intelligent application to parse emails, scrape websites, generate AI content, and send personalized HTML emails.",
# # # #     version="1.0.0"
# # # # )

# # # # # Initialize components (LLMProcessor might take time to load model)
# # # # # Model name can be configured via an environment variable if needed
# # # # HUGGING_FACE_MODEL_NAME = os.getenv("HUGGING_FACE_MODEL_NAME", "distilgpt2") # Default to distilgpt2
# # # # HUGGING_FACE_API_KEY = os.getenv("HUGGING_FACE_API_KEY") # Optional, for Hugging Face Inference API or private models

# # # # llm_processor = LLMProcessor(model_name=HUGGING_FACE_MODEL_NAME, api_key=HUGGING_FACE_API_KEY)
# # # # email_generator = EmailGenerator()

# # # # # SMTP Configuration from environment variables
# # # # SMTP_SERVER = os.getenv("SMTP_SERVER")
# # # # SMTP_PORT = int(os.getenv("SMTP_PORT", 587)) # Default to 587 if not set
# # # # SMTP_USERNAME = os.getenv("SMTP_USERNAME")
# # # # SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
# # # # SENDER_EMAIL = os.getenv("SENDER_EMAIL")
# # # # REPLY_TO_EMAIL = os.getenv("REPLY_TO_EMAIL")
# # # # TEST_RECEIVER_EMAIL = os.getenv("TEST_RECEIVER_EMAIL") # For testing or if a fixed recipient is desired

# # # # # Validate essential environment variables
# # # # if not all([SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL]):
# # # #     print("Warning: Missing one or more critical SMTP/email environment variables. Please check your .env file.")
# # # #     print("Required: SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL")
# # # #     email_sender = None # Prevent instantiation if config is incomplete
# # # # else:
# # # #     email_sender = EmailSender(SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD)


# # # # class EmailProcessRequest(BaseModel):
# # # #     email: str

# # # # @app.post("/process-email")
# # # # async def process_email_endpoint(request: EmailProcessRequest):
# # # #     """
# # # #     Processes a single email address:
# # # #     1. Parses the email to extract name and domain.
# # # #     2. Scrapes the associated company website.
# # # #     3. Uses an LLM to generate business summary and improvement tips.
# # # #     4. Generates a personalized HTML email.
# # # #     5. Sends the email via SMTP.
# # # #     """
# # # #     email_address = request.email
# # # #     print(f"Received request to process email: {email_address}")

# # # #     # 1. Parse Email
# # # #     parsed_info = parse_email(email_address)
# # # #     if not parsed_info["domain"] or not parsed_info["first_name"]:
# # # #         raise HTTPException(
# # # #             status_code=400,
# # # #             detail="Could not parse email address or extract domain/name. Please provide a valid email format."
# # # #         )

# # # #     recipient_name = parsed_info["first_name"]
# # # #     company_domain = parsed_info["domain"]
# # # #     # For email sending, the recipient will be a test email or the extracted email
# # # #     # For this example, we will send to TEST_RECEIVER_EMAIL for safety/testing,
# # # #     # but in a real app, you might send to `email_address` or a list of recipients.
# # # #     actual_recipient_email = TEST_RECEIVER_EMAIL # You can change this to `email_address` if you want to send directly
# # # #     company_name = parsed_info["full_name"].split(" ")[-1] if parsed_info["full_name"] else company_domain.split('.')[0].capitalize()


# # # #     print(f"Parsed Info: Name='{recipient_name}', Domain='{company_domain}', Company='{company_name}'")

# # # #     # 2. Web Scrape
# # # #     target_url = get_full_url(company_domain)
# # # #     print(f"Attempting to scrape URL: {target_url}")
# # # #     scraped_data = scrape_website_content(target_url)

# # # #     if scraped_data["error"]:
# # # #         print(f"Scraping error: {scraped_data['error']}")
# # # #         # We can still proceed with generic content if scraping fails, or raise an error
# # # #         # For now, let's allow it to proceed with empty scraped content
# # # #         scraped_content = ""
# # # #     else:
# # # #         # Combine relevant scraped data for LLM processing
# # # #         scraped_content = f"Title: {scraped_data['title'] or ''}\n" \
# # # #                           f"Description: {scraped_data['meta_description'] or ''}\n" \
# # # #                           f"Content: {scraped_data['main_content'] or ''}"
# # # #         print(f"Scraped content length: {len(scraped_content)} characters.")

# # # #     # 3. LLM Integration
# # # #     if not llm_processor.generator:
# # # #         raise HTTPException(status_code=500, detail="LLM model not initialized. Check server logs.")

# # # #     print("Generating business summary with LLM...")
# # # #     business_summary = llm_processor.summarize_business(company_name, scraped_content)
# # # #     print(f"Business Summary: {business_summary}")

# # # #     print("Generating improvement tips with LLM...")
# # # #     improvement_tips = llm_processor.generate_improvement_tips(company_name, business_summary, scraped_content)
# # # #     print(f"Improvement Tips: {improvement_tips}")

# # # #     print("Generating email body with LLM...")
# # # #     email_body = llm_processor.generate_email_body(recipient_name, company_name, business_summary, improvement_tips)
# # # #     print(f"Email Body (first 200 chars): {email_body[:200]}...")


# # # #     # 4. Generate HTML Email
# # # #     email_data = {
# # # #         "recipient_name": recipient_name,
# # # #         "company_name": company_name,
# # # #         "business_summary": business_summary,
# # # #         "improvement_tips": improvement_tips,
# # # #         "email_body": email_body,
# # # #         "sender_name": "Dave Chatpar", # Hardcoded as per assignment or configure in .env
# # # #         "sender_company": "NetWit",    # Hardcoded as per assignment or configure in .env
# # # #         "reply_to_email": REPLY_TO_EMAIL
# # # #     }
# # # #     html_email_content = email_generator.generate_email_html(email_data)
# # # #     print("HTML email content generated.")

# # # #     # 5. Send Email
# # # #     if not email_sender:
# # # #         raise HTTPException(status_code=500, detail="Email sender not configured. Check .env variables.")

# # # #     email_subject = f"Personalized Insights for Your Business: {company_name}"
# # # #     email_sent_success = email_sender.send_email(
# # # #         sender_email=SENDER_EMAIL,
# # # #         recipient_email=actual_recipient_email, # This will be the TEST_RECEIVER_EMAIL
# # # #         reply_to_email=REPLY_TO_EMAIL,
# # # #         subject=email_subject,
# # # #         html_content=html_email_content
# # # #     )

# # # #     if email_sent_success:
# # # #         return {
# # # #             "status": "success",
# # # #             "message": "Email processed and sent successfully!",
# # # #             "parsed_info": parsed_info,
# # # #             "scraped_data_summary": {
# # # #                 "title": scraped_data["title"],
# # # #                 "meta_description": scraped_data["meta_description"],
# # # #                 "content_length": len(scraped_data["main_content"])
# # # #             },
# # # #             "llm_generated_content": {
# # # #                 "business_summary": business_summary,
# # # #                 "improvement_tips": improvement_tips,
# # # #                 "email_body_preview": email_body[:150] + "..." # Truncate for response
# # # #             },
# # # #             "email_sent_to": actual_recipient_email
# # # #         }
# # # #     else:
# # # #         raise HTTPException(status_code=500, detail="Failed to send email. Check server logs for details.")

# # # # @app.get("/")
# # # # async def root():
# # # #     return {"message": "Email Automation API is running! Visit /docs for API documentation."}



# # # from fastapi import FastAPI, HTTPException
# # # from pydantic import BaseModel
# # # from dotenv import load_dotenv
# # # import os

# # # # Import modules from our app directory
# # # from app.email_parser import parse_email
# # # from app.web_scraper import get_full_url, scrape_website_content
# # # from app.llm_processor import LLMProcessor
# # # from app.email_generator import EmailGenerator
# # # from app.email_sender import EmailSender

# # # # Load environment variables from .env file
# # # load_dotenv()

# # # app = FastAPI(
# # #     title="Email Automation API",
# # #     description="An intelligent application to parse emails, scrape websites, generate AI content, and send personalized HTML emails.",
# # #     version="1.0.0"
# # # )

# # # # Initialize components (LLMProcessor might take time to load model)
# # # # Model name can be configured via an environment variable if needed
# # # HUGGING_FACE_MODEL_NAME = os.getenv("HUGGING_FACE_MODEL_NAME", "distilgpt2") # Default to distilgpt2
# # # HUGGING_FACE_API_KEY = os.getenv("HUGGING_FACE_API_KEY") # Optional, for Hugging Face Inference API or private models

# # # llm_processor = LLMProcessor(model_name=HUGGING_FACE_MODEL_NAME, api_key=HUGGING_FACE_API_KEY)
# # # email_generator = EmailGenerator()

# # # # SMTP Configuration from environment variables
# # # SMTP_SERVER = os.getenv("SMTP_SERVER")
# # # SMTP_PORT = int(os.getenv("SMTP_PORT", 587)) # Default to 587 if not set
# # # SMTP_USERNAME = os.getenv("SMTP_USERNAME")
# # # SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
# # # SENDER_EMAIL = os.getenv("SENDER_EMAIL")
# # # REPLY_TO_EMAIL = os.getenv("REPLY_TO_EMAIL")
# # # TEST_RECEIVER_EMAIL = os.getenv("TEST_RECEIVER_EMAIL") # For testing or if a fixed recipient is desired

# # # # Validate essential environment variables
# # # if not all([SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL]):
# # #     print("Warning: Missing one or more critical SMTP/email environment variables. Please check your .env file.")
# # #     print("Required: SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL")
# # #     email_sender = None # Prevent instantiation if config is incomplete
# # # else:
# # #     email_sender = EmailSender(SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD)


# # # class EmailProcessRequest(BaseModel):
# # #     email: str

# # # @app.post("/process-email")
# # # async def process_email_endpoint(request: EmailProcessRequest):
# # #     """
# # #     Processes a single email address:
# # #     1. Parses the email to extract name and domain.
# # #     2. Scrapes the associated company website.
# # #     3. Uses an LLM to generate business summary and improvement tips.
# # #     4. Generates a personalized HTML email.
# # #     5. Sends the email via SMTP.
# # #     """
# # #     email_address = request.email
# # #     print(f"Received request to process email: {email_address}")

# # #     # 1. Parse Email
# # #     parsed_info = parse_email(email_address)
# # #     if not parsed_info["domain"] or not parsed_info["first_name"]:
# # #         raise HTTPException(
# # #             status_code=400,
# # #             detail="Could not parse email address or extract domain/name. Please provide a valid email format."
# # #         )

# # #     recipient_name = parsed_info["first_name"]
# # #     company_domain = parsed_info["domain"]
# # #     # For email sending, the recipient will be a test email or the extracted email
# # #     # For this example, we will send to TEST_RECEIVER_EMAIL for safety/testing,
# # #     # but in a real app, you might send to `email_address` or a list of recipients.
# # #     actual_recipient_email = TEST_RECEIVER_EMAIL # You can change this to `email_address` if you want to send directly
# # #     company_name = parsed_info["full_name"].split(" ")[-1] if parsed_info["full_name"] else company_domain.split('.')[0].capitalize()


# # #     print(f"Parsed Info: Name='{recipient_name}', Domain='{company_domain}', Company='{company_name}'")

# # #     # 2. Web Scrape
# # #     target_url = get_full_url(company_domain)
# # #     print(f"Attempting to scrape URL: {target_url}")
# # #     scraped_data = scrape_website_content(target_url)

# # #     # Initialize with default/empty values
# # #     scraped_title = None
# # #     scraped_meta_description = None
# # #     scraped_main_content = "" # Default to empty string for LLM input

# # #     if scraped_data["error"]:
# # #         print(f"Scraping error encountered, proceeding with minimal content: {scraped_data['error']}")
# # #         # In this case, scraped_title, scraped_meta_description, scraped_main_content remain their initial values
# # #     else:
# # #         # Assign values if scraping was successful
# # #         scraped_title = scraped_data['title']
# # #         scraped_meta_description = scraped_data['meta_description']
# # #         scraped_main_content = scraped_data['main_content'] or "" # Ensure it's a string for len() and LLM input

# # #     # Combine relevant scraped data for LLM processing
# # #     # Ensure all parts are strings for concatenation
# # #     llm_input_content = f"Title: {scraped_title or ''}\n" \
# # #                         f"Description: {scraped_meta_description or ''}\n" \
# # #                         f"Content: {scraped_main_content or ''}"

# # #     print(f"Scraped content length for LLM: {len(llm_input_content)} characters.")


# # #     # 3. LLM Integration
# # #     if not llm_processor.generator:
# # #         raise HTTPException(status_code=500, detail="LLM model not initialized. Check server logs.")

# # #     print("Generating business summary with LLM...")
# # #     business_summary = llm_processor.summarize_business(company_name, llm_input_content)
# # #     print(f"Business Summary: {business_summary}")

# # #     print("Generating improvement tips with LLM...")
# # #     improvement_tips = llm_processor.generate_improvement_tips(company_name, business_summary, llm_input_content)
# # #     print(f"Improvement Tips: {improvement_tips}")

# # #     print("Generating email body with LLM...")
# # #     email_body = llm_processor.generate_email_body(recipient_name, company_name, business_summary, improvement_tips)
# # #     print(f"Email Body (first 200 chars): {email_body[:200]}...")


# # #     # 4. Generate HTML Email
# # #     email_data = {
# # #         "recipient_name": recipient_name,
# # #         "company_name": company_name,
# # #         "business_summary": business_summary,
# # #         "improvement_tips": improvement_tips,
# # #         "email_body": email_body,
# # #         "sender_name": "Dave Chatpar", # Hardcoded as per assignment or configure in .env
# # #         "sender_company": "NetWit",    # Hardcoded as per assignment or configure in .env
# # #         "reply_to_email": REPLY_TO_EMAIL
# # #     }
# # #     html_email_content = email_generator.generate_email_html(email_data)
# # #     print("HTML email content generated.")

# # #     # 5. Send Email
# # #     if not email_sender:
# # #         raise HTTPException(status_code=500, detail="Email sender not configured. Check .env variables.")

# # #     email_subject = f"Personalized Insights for Your Business: {company_name}"
# # #     email_sent_success = email_sender.send_email(
# # #         sender_email=SENDER_EMAIL,
# # #         recipient_email=actual_recipient_email, # This will be the TEST_RECEIVER_EMAIL
# # #         reply_to_email=REPLY_TO_EMAIL,
# # #         subject=email_subject,
# # #         html_content=html_email_content
# # #     )

# # #     if email_sent_success:
# # #         return {
# # #             "status": "success",
# # #             "message": "Email processed and sent successfully!",
# # #             "parsed_info": parsed_info,
# # #             "scraped_data_summary": {
# # #                 # Safely access values, providing None or 0 as fallback
# # #                 "title": scraped_title,
# # #                 "meta_description": scraped_meta_description,
# # #                 "content_length": len(scraped_main_content) # This will now be len("") if content was None
# # #             },
# # #             "llm_generated_content": {
# # #                 "business_summary": business_summary,
# # #                 "improvement_tips": improvement_tips,
# # #                 "email_body_preview": email_body[:150] + "..." # Truncate for response
# # #             },
# # #             "email_sent_to": actual_recipient_email
# # #         }
# # #     else:
# # #         raise HTTPException(status_code=500, detail="Failed to send email. Check server logs for details.")

# # # @app.get("/")
# # # async def root():
# # #     return {"message": "Email Automation API is running! Visit /docs for API documentation."}


# # from fastapi import FastAPI, HTTPException, UploadFile, File, Request
# # from fastapi.responses import HTMLResponse
# # from fastapi.templating import Jinja2Templates
# # from pydantic import BaseModel
# # from dotenv import load_dotenv
# # import os
# # import io # For reading file contents
# # import csv # For CSV parsing
# # import re # Added for extract_emails_from_file_content

# # # Try importing openpyxl for Excel, handle gracefully if not present
# # try:
# #     import openpyxl
# #     EXCEL_SUPPORT = True
# # except ImportError:
# #     print("Warning: openpyxl not found. Excel file uploads will not be processed.")
# #     EXCEL_SUPPORT = False


# # # Import modules from our app directory
# # from app.email_parser import parse_email
# # from app.web_scraper import get_full_url, scrape_website_content
# # from app.llm_processor import LLMProcessor
# # from app.email_generator import EmailGenerator
# # from app.email_sender import EmailSender

# # # Load environment variables from .env file
# # load_dotenv()

# # app = FastAPI(
# #     title="Email Automation API",
# #     description="An intelligent application to parse emails, scrape websites, generate AI content, and send personalized HTML emails.",
# #     version="1.0.0"
# # )

# # # Configure Jinja2Templates to serve HTML files
# # templates = Jinja2Templates(directory="app/templates")

# # # Initialize components (LLMProcessor might take time to load model)
# # # Model name can be configured via an environment variable if needed
# # HUGGING_FACE_MODEL_NAME = os.getenv("HUGGING_FACE_MODEL_NAME", "distilgpt2") # Default to distilgpt2
# # HUGGING_FACE_API_KEY = os.getenv("HUGGING_FACE_API_KEY") # Optional, for Hugging Face Inference API or private models

# # llm_processor = LLMProcessor(model_name=HUGGING_FACE_MODEL_NAME, api_key=HUGGING_FACE_API_KEY)
# # email_generator = EmailGenerator()

# # # SMTP Configuration from environment variables
# # SMTP_SERVER = os.getenv("SMTP_SERVER")
# # SMTP_PORT = int(os.getenv("SMTP_PORT", 587)) # Default to 587 if not set
# # SMTP_USERNAME = os.getenv("SMTP_USERNAME")
# # SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
# # SENDER_EMAIL = os.getenv("SENDER_EMAIL")
# # REPLY_TO_EMAIL = os.getenv("REPLY_TO_EMAIL")
# # TEST_RECEIVER_EMAIL = os.getenv("TEST_RECEIVER_EMAIL") # For testing or if a fixed recipient is desired

# # # Validate essential environment variables
# # if not all([SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL]):
# #     print("Warning: Missing one or more critical SMTP/email environment variables. Please check your .env file.")
# #     print("Required: SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL")
# #     email_sender = None # Prevent instantiation if config is incomplete
# # else:
# #     email_sender = EmailSender(SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD)


# # class EmailProcessRequest(BaseModel):
# #     email: str

# # # Helper function to extract emails from file content
# # def extract_emails_from_file_content(file_content: bytes, filename: str) -> list[str]:
# #     """
# #     Extracts email addresses from the content of an uploaded file.
# #     Supports .txt, .csv, and .xlsx (if openpyxl is installed).
# #     Assumes emails are either line-separated or comma-separated in text/csv.
# #     For Excel, it reads the first column of the first sheet.
# #     """
# #     emails = []
# #     decoded_content = None

# #     if filename.endswith(('.txt', '.csv')):
# #         try:
# #             decoded_content = file_content.decode('utf-8')
# #         except UnicodeDecodeError:
# #             decoded_content = file_content.decode('latin-1') # Fallback for other encodings

# #         # Try to parse as CSV first, then fallback to line-by-line
# #         reader = csv.reader(io.StringIO(decoded_content))
# #         for row in reader:
# #             for item in row:
# #                 # A simple regex to detect email-like strings
# #                 email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', item)
# #                 if email_match:
# #                     emails.append(email_match.group(0))
# #         if not emails: # If CSV parsing didn't find emails, try line by line
# #             for line in decoded_content.splitlines():
# #                 email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', line)
# #                 if email_match:
# #                     emails.append(email_match.group(0))

# #     elif filename.endswith(('.xlsx', '.xls')):
# #         if not EXCEL_SUPPORT:
# #             print("Excel file detected but openpyxl is not installed. Cannot process.")
# #             raise HTTPException(status_code=400, detail="Excel file processing requires 'openpyxl'. Please install it (`pip install openpyxl`).")
        
# #         try:
# #             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
# #             sheet = workbook.active
# #             for row in sheet.iter_rows():
# #                 for cell in row:
# #                     if cell.value and isinstance(cell.value, str):
# #                         email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', cell.value)
# #                         if email_match:
# #                             emails.append(email_match.group(0))
# #         except Exception as e:
# #             print(f"Error processing Excel file: {e}")
# #             raise HTTPException(status_code=400, detail=f"Error reading Excel file: {e}. Ensure it's a valid Excel format.")
# #     else:
# #         raise HTTPException(status_code=400, detail="Unsupported file type. Please upload .txt, .csv, or .xlsx files.")

# #     # Deduplicate emails and return
# #     return list(dict.fromkeys(emails))


# # # New UI Endpoint to serve the HTML file
# # @app.get("/", response_class=HTMLResponse)
# # async def read_root(request: Request):
# #     """
# #     Serves the main HTML file for file uploads.
# #     """
# #     return templates.TemplateResponse("upload.html", {"request": request})

# # # New File Upload Endpoint
# # @app.post("/uploadfile")
# # async def create_upload_file(file: UploadFile = File(...)):
# #     """
# #     Handles uploaded email list files, extracts emails, and processes each one.
# #     """
# #     print(f"Received file upload: {file.filename}, Content-Type: {file.content_type}")

# #     if not (file.filename.endswith(('.txt', '.csv', '.xlsx', '.xls'))):
# #         raise HTTPException(status_code=400, detail="Invalid file type. Only .txt, .csv, .xlsx, and .xls files are supported.")

# #     file_contents = await file.read()
# #     emails_to_process = []
# #     processed_count = 0
# #     failed_count = 0
# #     errors = []

# #     try:
# #         emails_to_process = extract_emails_from_file_content(file_contents, file.filename)
# #         if not emails_to_process:
# #             raise HTTPException(status_code=400, detail="No valid email addresses found in the uploaded file.")
        
# #         print(f"Found {len(emails_to_process)} emails to process.")

# #         for email_address in emails_to_process:
# #             print(f"Processing email from uploaded file: {email_address}")
# #             try:
# #                 # --- Existing Logic from /process-email endpoint, modified slightly ---
# #                 parsed_info = parse_email(email_address)
# #                 if not parsed_info["domain"] or not parsed_info["first_name"]:
# #                     print(f"Skipping email '{email_address}': Could not parse name/domain.")
# #                     failed_count += 1
# #                     errors.append(f"Failed to parse or missing info for email: {email_address}")
# #                     continue

# #                 recipient_name = parsed_info["first_name"]
# #                 company_domain = parsed_info["domain"]
# #                 actual_recipient_email = TEST_RECEIVER_EMAIL # Send to test recipient for all, or change to email_address
# #                 company_name = parsed_info["full_name"].split(" ")[-1] if parsed_info["full_name"] else company_domain.split('.')[0].capitalize()

# #                 target_url = get_full_url(company_domain)
# #                 scraped_data = scrape_website_content(target_url)

# #                 scraped_title = None
# #                 scraped_meta_description = None
# #                 scraped_main_content = ""

# #                 if scraped_data["error"]:
# #                     print(f"Scraping error for {target_url}: {scraped_data['error']}")
# #                     errors.append(f"Scraping failed for {company_domain}: {scraped_data['error']}")
# #                 else:
# #                     scraped_title = scraped_data['title']
# #                     scraped_meta_description = scraped_data['meta_description']
# #                     scraped_main_content = scraped_data['main_content'] or ""

# #                 llm_input_content = f"Title: {scraped_title or ''}\n" \
# #                                     f"Description: {scraped_meta_description or ''}\n" \
# #                                     f"Content: {scraped_main_content or ''}"
                
# #                 if not llm_processor.generator:
# #                     print("LLM not initialized, skipping content generation.")
# #                     errors.append(f"LLM not available for {email_address}.")
# #                     failed_count += 1
# #                     continue

# #                 business_summary = llm_processor.summarize_business(company_name, llm_input_content)
# #                 improvement_tips = llm_processor.generate_improvement_tips(company_name, business_summary, llm_input_content)
# #                 email_body = llm_processor.generate_email_body(recipient_name, company_name, business_summary, improvement_tips)

# #                 email_data = {
# #                     "recipient_name": recipient_name,
# #                     "company_name": company_name,
# #                     "business_summary": business_summary,
# #                     "improvement_tips": improvement_tips,
# #                     "email_body": email_body,
# #                     "sender_name": "Dave Chatpar",
# #                     "sender_company": "NetWit",
# #                     "reply_to_email": REPLY_TO_EMAIL
# #                 }
# #                 html_email_content = email_generator.generate_email_html(email_data)

# #                 if not email_sender:
# #                     print("Email sender not configured, skipping email sending.")
# #                     errors.append(f"Email sender not configured for {email_address}.")
# #                     failed_count += 1
# #                     continue

# #                 email_subject = f"Personalized Insights for Your Business: {company_name}"
# #                 email_sent_success = email_sender.send_email(
# #                     sender_email=SENDER_EMAIL,
# #                     recipient_email=actual_recipient_email,
# #                     reply_to_email=REPLY_TO_EMAIL,
# #                     subject=email_subject,
# #                     html_content=html_email_content
# #                 )

# #                 if email_sent_success:
# #                     processed_count += 1
# #                 else:
# #                     failed_count += 1
# #                     errors.append(f"Failed to send email to {actual_recipient_email} for {email_address}.")

# #             except Exception as e:
# #                 print(f"Error processing individual email '{email_address}': {e}")
# #                 failed_count += 1
# #                 errors.append(f"Processing failed for email {email_address}: {e}")

# #         # --- End of Existing Logic ---

# #         message = f"Successfully processed {processed_count} emails. "
# #         if failed_count > 0:
# #             message += f"Failed to process {failed_count} emails. See details below."
# #         elif processed_count == 0 and failed_count == 0 and len(emails_to_process) > 0:
# #              message = "No emails were processed. Please check your file content and server logs for errors."
# #         elif len(emails_to_process) == 0:
# #             message = "The uploaded file did not contain any email addresses."


# #         return {
# #             "message": message,
# #             "total_emails_in_file": len(emails_to_process),
# #             "successfully_processed": processed_count,
# #             "failed_to_process": failed_count,
# #             "errors": errors
# #         }

# #     except HTTPException as e:
# #         raise e # Re-raise FastAPI HTTPExceptions
# #     except Exception as e:
# #         print(f"Overall file processing error: {e}")
# #         raise HTTPException(status_code=500, detail=f"An error occurred during file processing: {e}")

# # # Keep the original /process-email endpoint for single email processing (optional, but good for API consistency)
# # @app.post("/process-email")
# # async def process_single_email_endpoint(request: EmailProcessRequest):
# #     """
# #     Processes a single email address:
# #     1. Parses the email to extract name and domain.
# #     2. Scrapes the associated company website.
# #     3. Uses an LLM to generate business summary and improvement tips.
# #     4. Generates a personalized HTML email.
# #     5. Sends the email via SMTP.
# #     """
# #     email_address = request.email
# #     print(f"Received request to process single email: {email_address}")

# #     # 1. Parse Email
# #     parsed_info = parse_email(email_address)
# #     if not parsed_info["domain"] or not parsed_info["first_name"]:
# #         raise HTTPException(
# #             status_code=400,
# #             detail="Could not parse email address or extract domain/name. Please provide a valid email format."
# #         )

# #     recipient_name = parsed_info["first_name"]
# #     company_domain = parsed_info["domain"]
# #     actual_recipient_email = TEST_RECEIVER_EMAIL # Send to test recipient for all, or change to email_address
# #     company_name = parsed_info["full_name"].split(" ")[-1] if parsed_info["full_name"] else company_domain.split('.')[0].capitalize()


# #     print(f"Parsed Info: Name='{recipient_name}', Domain='{company_domain}', Company='{company_name}'")

# #     # 2. Web Scraping
# #     target_url = get_full_url(company_domain)
# #     print(f"Attempting to scrape URL: {target_url}")
# #     scraped_data = scrape_website_content(target_url)

# #     scraped_title = None
# #     scraped_meta_description = None
# #     scraped_main_content = ""

# #     if scraped_data["error"]:
# #         print(f"Scraping error encountered, proceeding with minimal content: {scraped_data['error']}")
# #     else:
# #         scraped_title = scraped_data['title']
# #         scraped_meta_description = scraped_data['meta_description']
# #         scraped_main_content = scraped_data['main_content'] or ""

# #     llm_input_content = f"Title: {scraped_title or ''}\n" \
# #                         f"Description: {scraped_meta_description or ''}\n" \
# #                         f"Content: {scraped_main_content or ''}"

# #     print(f"Scraped content length for LLM: {len(llm_input_content)} characters.")

# #     # 3. LLM Integration
# #     if not llm_processor.generator:
# #         raise HTTPException(status_code=500, detail="LLM model not initialized. Check server logs.")

# #     print("Generating business summary with LLM...")
# #     business_summary = llm_processor.summarize_business(company_name, llm_input_content)
# #     print(f"Business Summary: {business_summary}")

# #     print("Generating improvement tips with LLM...")
# #     improvement_tips = llm_processor.generate_improvement_tips(company_name, business_summary, llm_input_content)
# #     print(f"Improvement Tips: {improvement_tips}")

# #     print("Generating email body with LLM...")
# #     email_body = llm_processor.generate_email_body(recipient_name, company_name, business_summary, improvement_tips)
# #     print(f"Email Body (first 200 chars): {email_body[:200]}...")


# #     # 4. Generate HTML Email
# #     email_data = {
# #         "recipient_name": recipient_name,
# #         "company_name": company_name,
# #         "business_summary": business_summary,
# #         "improvement_tips": improvement_tips,
# #         "email_body": email_body,
# #         "sender_name": "Dave Chatpar",
# #         "sender_company": "NetWit",
# #         "reply_to_email": REPLY_TO_EMAIL
# #     }
# #     html_email_content = email_generator.generate_email_html(email_data)
# #     print("HTML email content generated.")

# #     # 5. Send Email
# #     if not email_sender:
# #         raise HTTPException(status_code=500, detail="Email sender not configured. Check .env variables.")

# #     email_subject = f"Personalized Insights for Your Business: {company_name}"
# #     email_sent_success = email_sender.send_email(
# #         sender_email=SENDER_EMAIL,
# #         recipient_email=actual_recipient_email,
# #         reply_to_email=REPLY_TO_EMAIL,
# #         subject=email_subject,
# #         html_content=html_email_content
# #     )

# #     if email_sent_success:
# #         return {
# #             "status": "success",
# #             "message": "Email processed and sent successfully!",
# #             "parsed_info": parsed_info,
# #             "scraped_data_summary": {
# #                 "title": scraped_title,
# #                 "meta_description": scraped_meta_description,
# #                 "content_length": len(scraped_main_content)
# #             },
# #             "llm_generated_content": {
# #                 "business_summary": business_summary,
# #                 "improvement_tips": improvement_tips,
# #                 "email_body_preview": email_body[:150] + "..."
# #             },
# #             "email_sent_to": actual_recipient_email
# #         }
# #     else:
# #         raise HTTPException(status_code=500, detail="Failed to send email. Check server logs for details.")



# from fastapi import FastAPI, HTTPException, UploadFile, File, Request
# from fastapi.responses import HTMLResponse
# from fastapi.templating import Jinja2Templates
# from pydantic import BaseModel
# from dotenv import load_dotenv
# import os
# import io # For reading file contents
# import csv # For CSV parsing
# import re # Added for extract_emails_from_file_content

# # Try importing openpyxl for Excel, handle gracefully if not present
# try:
#     import openpyxl
#     EXCEL_SUPPORT = True
# except ImportError:
#     print("Warning: openpyxl not found. Excel file uploads will not be processed.")
#     EXCEL_SUPPORT = False


# # Import modules from our app directory
# from app.email_parser import parse_email
# from app.web_scraper import get_full_url, scrape_website_content
# from app.llm_processor import LLMProcessor
# from app.email_generator import EmailGenerator
# from app.email_sender import EmailSender
# from app.database import DatabaseManager # NEW: Import DatabaseManager

# # Load environment variables from .env file
# load_dotenv()

# app = FastAPI(
#     title="Email Automation API",
#     description="An intelligent application to parse emails, scrape websites, generate AI content, and send personalized HTML emails.",
#     version="1.0.0"
# )

# # Configure Jinja2Templates to serve HTML files
# templates = Jinja2Templates(directory="app/templates")

# # Initialize components (LLMProcessor might take time to load model)
# # Model name can be configured via an environment variable if needed
# HUGGING_FACE_MODEL_NAME = os.getenv("HUGGING_FACE_MODEL_NAME", "distilgpt2") # Default to distilgpt2
# HUGGING_FACE_API_KEY = os.getenv("HUGGING_FACE_API_KEY") # Optional, for Hugging Face Inference API or private models

# llm_processor = LLMProcessor(model_name=HUGGING_FACE_MODEL_NAME, api_key=HUGGING_FACE_API_KEY)
# email_generator = EmailGenerator()

# # NEW: Initialize DatabaseManager
# db_manager = DatabaseManager()


# # SMTP Configuration from environment variables
# SMTP_SERVER = os.getenv("SMTP_SERVER")
# SMTP_PORT = int(os.getenv("SMTP_PORT", 587)) # Default to 587 if not set
# SMTP_USERNAME = os.getenv("SMTP_USERNAME")
# SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
# SENDER_EMAIL = os.getenv("SENDER_EMAIL")
# REPLY_TO_EMAIL = os.getenv("REPLY_TO_EMAIL")
# TEST_RECEIVER_EMAIL = os.getenv("TEST_RECEIVER_EMAIL") # For testing or if a fixed recipient is desired

# # Validate essential environment variables
# if not all([SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL]):
#     print("Warning: Missing one or more critical SMTP/email environment variables. Please check your .env file.")
#     print("Required: SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL")
#     email_sender = None # Prevent instantiation if config is incomplete
# else:
#     email_sender = EmailSender(SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD)


# class EmailProcessRequest(BaseModel):
#     email: str

# # Helper function to extract emails from file content
# def extract_emails_from_file_content(file_content: bytes, filename: str) -> list[str]:
#     """
#     Extracts email addresses from the content of an uploaded file.
#     Supports .txt, .csv, and .xlsx (if openpyxl is installed).
#     Assumes emails are either line-separated or comma-separated in text/csv.
#     For Excel, it reads the first column of the first sheet.
#     """
#     emails = []
#     decoded_content = None

#     if filename.endswith(('.txt', '.csv')):
#         try:
#             decoded_content = file_content.decode('utf-8')
#         except UnicodeDecodeError:
#             decoded_content = file_content.decode('latin-1') # Fallback for other encodings

#         # Try to parse as CSV first, then fallback to line-by-line
#         reader = csv.reader(io.StringIO(decoded_content))
#         for row in reader:
#             for item in row:
#                 # A simple regex to detect email-like strings
#                 email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', item)
#                 if email_match:
#                     emails.append(email_match.group(0))
#         if not emails: # If CSV parsing didn't find emails, try line by line
#             for line in decoded_content.splitlines():
#                 email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', line)
#                 if email_match:
#                     emails.append(email_match.group(0))

#     elif filename.endswith(('.xlsx', '.xls')):
#         if not EXCEL_SUPPORT:
#             print("Excel file detected but openpyxl is not installed. Cannot process.")
#             raise HTTPException(status_code=400, detail="Excel file processing requires 'openpyxl'. Please install it (`pip install openpyxl`).")
        
#         try:
#             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
#             sheet = workbook.active
#             for row in sheet.iter_rows():
#                 for cell in row:
#                     if cell.value and isinstance(cell.value, str):
#                         email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', cell.value)
#                         if email_match:
#                             emails.append(email_match.group(0))
#         except Exception as e:
#             print(f"Error processing Excel file: {e}")
#             raise HTTPException(status_code=400, detail=f"Error reading Excel file: {e}. Ensure it's a valid Excel format.")
#     else:
#         raise HTTPException(status_code=400, detail="Unsupported file type. Please upload .txt, .csv, or .xlsx files.")

#     # Deduplicate emails and return
#     return list(dict.fromkeys(emails))


# # UI Endpoint to serve the HTML file
# @app.get("/", response_class=HTMLResponse)
# async def read_root(request: Request):
#     """
#     Serves the main HTML file for file uploads.
#     """
#     return templates.TemplateResponse("upload.html", {"request": request})

# # File Upload Endpoint
# @app.post("/uploadfile")
# async def create_upload_file(file: UploadFile = File(...)):
#     """
#     Handles uploaded email list files, extracts emails, and processes each one.
#     """
#     print(f"Received file upload: {file.filename}, Content-Type: {file.content_type}")

#     if not (file.filename.endswith(('.txt', '.csv', '.xlsx', '.xls'))):
#         raise HTTPException(status_code=400, detail="Invalid file type. Only .txt, .csv, .xlsx, and .xls files are supported.")

#     file_contents = await file.read()
#     emails_to_process = []
#     processed_count = 0
#     failed_count = 0
#     errors = []

#     try:
#         emails_to_process = extract_emails_from_file_content(file_contents, file.filename)
#         if not emails_to_process:
#             raise HTTPException(status_code=400, detail="No valid email addresses found in the uploaded file.")
        
#         print(f"Found {len(emails_to_process)} emails to process.")

#         for email_address in emails_to_process:
#             print(f"Processing email from uploaded file: {email_address}")
#             current_email_status = "failed" # Default status
#             current_error_message = None
#             email_subject_to_log = "N/A" # Default for logging if not generated
#             company_domain_to_log = "N/A"
#             recipient_email_to_log = TEST_RECEIVER_EMAIL

#             try:
#                 # --- Existing Logic from /process-email endpoint, modified slightly ---
#                 parsed_info = parse_email(email_address)
#                 if not parsed_info["domain"] or not parsed_info["first_name"]:
#                     print(f"Skipping email '{email_address}': Could not parse name/domain.")
#                     current_error_message = f"Could not parse name/domain for {email_address}."
#                     failed_count += 1
#                     errors.append(current_error_message)
#                     db_manager.log_email_processing(
#                         original_email=email_address,
#                         recipient_email_sent_to=recipient_email_to_log,
#                         company_domain=parsed_info["domain"] if parsed_info["domain"] else "N/A",
#                         email_subject="Parsing Failed",
#                         status="failed",
#                         error_message=current_error_message
#                     )
#                     continue

#                 recipient_name = parsed_info["first_name"]
#                 company_domain = parsed_info["domain"]
#                 company_domain_to_log = company_domain # Update for logging
                
#                 # actual_recipient_email = TEST_RECEIVER_EMAIL # Already set above as recipient_email_to_log

#                 company_name = parsed_info["full_name"].split(" ")[-1] if parsed_info["full_name"] else company_domain.split('.')[0].capitalize()

#                 target_url = get_full_url(company_domain)
#                 scraped_data = scrape_website_content(target_url)

#                 scraped_title = None
#                 scraped_meta_description = None
#                 scraped_main_content = ""

#                 if scraped_data["error"]:
#                     print(f"Scraping error for {target_url}: {scraped_data['error']}")
#                     current_error_message = f"Scraping failed for {company_domain}: {scraped_data['error']}"
#                     errors.append(current_error_message)
#                 else:
#                     scraped_title = scraped_data['title']
#                     scraped_meta_description = scraped_data['meta_description']
#                     scraped_main_content = scraped_data['main_content'] or ""

#                 llm_input_content = f"Title: {scraped_title or ''}\n" \
#                                     f"Description: {scraped_meta_description or ''}\n" \
#                                     f"Content: {scraped_main_content or ''}"
                
#                 if not llm_processor.generator:
#                     print("LLM not initialized, skipping content generation.")
#                     current_error_message = f"LLM not available for {email_address}."
#                     errors.append(current_error_message)
#                     failed_count += 1
#                     db_manager.log_email_processing(
#                         original_email=email_address,
#                         recipient_email_sent_to=recipient_email_to_log,
#                         company_domain=company_domain_to_log,
#                         email_subject="LLM Init Failed",
#                         status="failed",
#                         error_message=current_error_message
#                     )
#                     continue

#                 business_summary = llm_processor.summarize_business(company_name, llm_input_content)
#                 improvement_tips = llm_processor.generate_improvement_tips(company_name, business_summary, llm_input_content)
#                 email_body = llm_processor.generate_email_body(recipient_name, company_name, business_summary, improvement_tips)

#                 email_data = {
#                     "recipient_name": recipient_name,
#                     "company_name": company_name,
#                     "business_summary": business_summary,
#                     "improvement_tips": improvement_tips,
#                     "email_body": email_body,
#                     "sender_name": "Dave Chatpar",
#                     "sender_company": "NetWit",
#                     "reply_to_email": REPLY_TO_EMAIL
#                 }
#                 html_email_content = email_generator.generate_email_html(email_data)

#                 email_subject = f"Personalized Insights for Your Business: {company_name}"
#                 email_subject_to_log = email_subject # Update for logging

#                 if not email_sender:
#                     print("Email sender not configured, skipping email sending.")
#                     current_error_message = f"Email sender not configured for {email_address}."
#                     errors.append(current_error_message)
#                     failed_count += 1
#                     db_manager.log_email_processing(
#                         original_email=email_address,
#                         recipient_email_sent_to=recipient_email_to_log,
#                         company_domain=company_domain_to_log,
#                         email_subject=email_subject_to_log,
#                         status="failed",
#                         error_message=current_error_message
#                     )
#                     continue

#                 email_sent_success = email_sender.send_email(
#                     sender_email=SENDER_EMAIL,
#                     recipient_email=recipient_email_to_log,
#                     reply_to_email=REPLY_TO_EMAIL,
#                     subject=email_subject,
#                     html_content=html_email_content
#                 )

#                 if email_sent_success:
#                     processed_count += 1
#                     current_email_status = "success"
#                 else:
#                     failed_count += 1
#                     current_email_status = "failed"
#                     current_error_message = f"Failed to send email to {recipient_email_to_log} for {email_address}. Check SMTP logs."
#                     errors.append(current_error_message)

#             except Exception as e:
#                 print(f"Error processing individual email '{email_address}': {e}")
#                 failed_count += 1
#                 current_email_status = "failed"
#                 current_error_message = f"Processing failed for email {email_address}: {e}"
#                 errors.append(current_error_message)
#             finally:
#                 # Log the outcome for EACH email, even if an exception occurred within the loop
#                 db_manager.log_email_processing(
#                     original_email=email_address,
#                     recipient_email_sent_to=recipient_email_to_log,
#                     company_domain=company_domain_to_log,
#                     email_subject=email_subject_to_log,
#                     status=current_email_status,
#                     error_message=current_error_message
#                 )

#         # --- End of Existing Logic ---

#         message = f"Successfully processed {processed_count} emails. "
#         if failed_count > 0:
#             message += f"Failed to process {failed_count} emails. See details below."
#         elif processed_count == 0 and failed_count == 0 and len(emails_to_process) > 0:
#              message = "No emails were processed. Please check your file content and server logs for errors."
#         elif len(emails_to_process) == 0:
#             message = "The uploaded file did not contain any valid email addresses."


#         return {
#             "message": message,
#             "total_emails_in_file": len(emails_to_process),
#             "successfully_processed": processed_count,
#             "failed_to_process": failed_count,
#             "errors": errors
#         }

#     except HTTPException as e:
#         raise e # Re-raise FastAPI HTTPExceptions
#     except Exception as e:
#         print(f"Overall file processing error: {e}")
#         raise HTTPException(status_code=500, detail=f"An error occurred during file processing: {e}")

# # Keep the original /process-email endpoint for single email processing (optional, but good for API consistency)
# @app.post("/process-email")
# async def process_single_email_endpoint(request: EmailProcessRequest):
#     """
#     Processes a single email address:
#     1. Parses the email to extract name and domain.
#     2. Scrapes the associated company website.
#     3. Uses an LLM to generate business summary and improvement tips.
#     4. Generates a personalized HTML email.
#     5. Sends the email via SMTP.
#     (This endpoint also logs to the database)
#     """
#     email_address = request.email
#     print(f"Received request to process single email: {email_address}")

#     current_email_status = "failed" # Default status for single email
#     current_error_message = None
#     email_subject_to_log = "N/A"
#     company_domain_to_log = "N/A"
#     recipient_email_to_log = TEST_RECEIVER_EMAIL # Will be used in finally block


#     try:
#         # 1. Parse Email
#         parsed_info = parse_email(email_address)
#         if not parsed_info["domain"] or not parsed_info["first_name"]:
#             current_error_message = "Could not parse email address or extract domain/name."
#             raise HTTPException(status_code=400, detail=current_error_message)

#         recipient_name = parsed_info["first_name"]
#         company_domain = parsed_info["domain"]
#         company_domain_to_log = company_domain
        
#         company_name = parsed_info["full_name"].split(" ")[-1] if parsed_info["full_name"] else company_domain.split('.')[0].capitalize()


#         print(f"Parsed Info: Name='{recipient_name}', Domain='{company_domain}', Company='{company_name}'")

#         # 2. Web Scraping
#         target_url = get_full_url(company_domain)
#         print(f"Attempting to scrape URL: {target_url}")
#         scraped_data = scrape_website_content(target_url)

#         scraped_title = None
#         scraped_meta_description = None
#         scraped_main_content = ""

#         if scraped_data["error"]:
#             print(f"Scraping error encountered, proceeding with minimal content: {scraped_data['error']}")
#             current_error_message = f"Scraping error: {scraped_data['error']}" # Log scraping errors
#         else:
#             scraped_title = scraped_data['title']
#             scraped_meta_description = scraped_data['meta_description']
#             scraped_main_content = scraped_data['main_content'] or ""

#         llm_input_content = f"Title: {scraped_title or ''}\n" \
#                             f"Description: {scraped_meta_description or ''}\n" \
#                             f"Content: {scraped_main_content or ''}"

#         print(f"Scraped content length for LLM: {len(llm_input_content)} characters.")

#         # 3. LLM Integration
#         if not llm_processor.generator:
#             current_error_message = "LLM model not initialized. Check server logs."
#             raise HTTPException(status_code=500, detail=current_error_message)

#         print("Generating business summary with LLM...")
#         business_summary = llm_processor.summarize_business(company_name, llm_input_content)
#         print(f"Business Summary: {business_summary}")

#         print("Generating improvement tips with LLM...")
#         improvement_tips = llm_processor.generate_improvement_tips(company_name, business_summary, llm_input_content)
#         print(f"Improvement Tips: {improvement_tips}")

#         print("Generating email body with LLM...")
#         email_body = llm_processor.generate_email_body(recipient_name, company_name, business_summary, improvement_tips)
#         print(f"Email Body (first 200 chars): {email_body[:200]}...")


#         # 4. Generate HTML Email
#         email_data = {
#             "recipient_name": recipient_name,
#             "company_name": company_name,
#             "business_summary": business_summary,
#             "improvement_tips": improvement_tips,
#             "email_body": email_body,
#             "sender_name": "Dave Chatpar",
#             "sender_company": "NetWit",
#             "reply_to_email": REPLY_TO_EMAIL
#         }
#         html_email_content = email_generator.generate_email_html(email_data)
#         print("HTML email content generated.")

#         email_subject = f"Personalized Insights for Your Business: {company_name}"
#         email_subject_to_log = email_subject # Update for logging

#         # 5. Send Email
#         if not email_sender:
#             current_error_message = "Email sender not configured. Check .env variables."
#             raise HTTPException(status_code=500, detail=current_error_message)

#         email_sent_success = email_sender.send_email(
#             sender_email=SENDER_EMAIL,
#             recipient_email=recipient_email_to_log,
#             reply_to_email=REPLY_TO_EMAIL,
#             subject=email_subject,
#             html_content=html_email_content
#         )

#         if email_sent_success:
#             current_email_status = "success"
#             return {
#                 "status": "success",
#                 "message": "Email processed and sent successfully!",
#                 "parsed_info": parsed_info,
#                 "scraped_data_summary": {
#                     "title": scraped_title,
#                     "meta_description": scraped_meta_description,
#                     "content_length": len(scraped_main_content)
#                 },
#                 "llm_generated_content": {
#                     "business_summary": business_summary,
#                     "improvement_tips": improvement_tips,
#                     "email_body_preview": email_body[:150] + "..."
#                 },
#                 "email_sent_to": recipient_email_to_log
#             }
#         else:
#             current_error_message = "Failed to send email. Check server logs for details."
#             raise HTTPException(status_code=500, detail=current_error_message)

#     except HTTPException as http_exc:
#         current_error_message = http_exc.detail
#         raise http_exc # Re-raise FastAPI HTTPExceptions
#     except Exception as e:
#         print(f"Overall processing error for single email '{email_address}': {e}")
#         current_error_message = f"An unexpected error occurred: {e}"
#         raise HTTPException(status_code=500, detail=current_error_message)
#     finally:
#         # Log the outcome for the single email, even if an exception occurred
#         db_manager.log_email_processing(
#             original_email=email_address,
#             recipient_email_sent_to=recipient_email_to_log,
#             company_domain=company_domain_to_log,
#             email_subject=email_subject_to_log,
#             status=current_email_status,
#             error_message=current_error_message
#         )

from fastapi import FastAPI, HTTPException, UploadFile, File, Request
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from dotenv import load_dotenv
import os
import io # For reading file contents
import csv # For CSV parsing
import re # Added for extract_emails_from_file_content

# Try importing openpyxl for Excel, handle gracefully if not present
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    print("Warning: openpyxl not found. Excel file uploads will not be processed.")
    EXCEL_SUPPORT = False


# Import modules from our app directory
from app.email_parser import parse_email
from app.web_scraper import get_full_url, scrape_website_content
from app.llm_processor import LLMProcessor
from app.email_generator import EmailGenerator
from app.email_sender import EmailSender
from app.database import DatabaseManager # NEW: Import DatabaseManager

# Load environment variables from .env file
load_dotenv()

app = FastAPI(
    title="Email Automation API",
    description="An intelligent application to parse emails, scrape websites, generate AI content, and send personalized HTML emails.",
    version="1.0.0"
)

# Configure Jinja2Templates to serve HTML files
templates = Jinja2Templates(directory="app/templates")

# Initialize components (LLMProcessor might take time to load model)
# Model name can be configured via an environment variable if needed
HUGGING_FACE_MODEL_NAME = os.getenv("HUGGING_FACE_MODEL_NAME", "distilgpt2") # Default to distilgpt2
HUGGING_FACE_API_KEY = os.getenv("HUGGING_FACE_API_KEY") # Optional, for Hugging Face Inference API or private models

llm_processor = LLMProcessor(model_name=HUGGING_FACE_MODEL_NAME, api_key=HUGGING_FACE_API_KEY)
email_generator = EmailGenerator()

# NEW: Initialize DatabaseManager
db_manager = DatabaseManager()


# SMTP Configuration from environment variables
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587)) # Default to 587 if not set
SMTP_USERNAME = os.getenv("SMTP_USERNAME")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
REPLY_TO_EMAIL = os.getenv("REPLY_TO_EMAIL")
TEST_RECEIVER_EMAIL = os.getenv("TEST_RECEIVER_EMAIL") # For testing or if a fixed recipient is desired

# Validate essential environment variables
if not all([SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL]):
    print("Warning: Missing one or more critical SMTP/email environment variables. Please check your .env file.")
    print("Required: SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SENDER_EMAIL, REPLY_TO_EMAIL, TEST_RECEIVER_EMAIL")
    email_sender = None # Prevent instantiation if config is incomplete
else:
    email_sender = EmailSender(SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD)


class EmailProcessRequest(BaseModel):
    email: str

# Helper function to extract emails from file content
def extract_emails_from_file_content(file_content: bytes, filename: str) -> list[str]:
    """
    Extracts email addresses from the content of an uploaded file.
    Supports .txt, .csv, and .xlsx (if openpyxl is installed).
    Assumes emails are either line-separated or comma-separated in text/csv.
    For Excel, it reads the first column of the first sheet.
    """
    emails = []
    decoded_content = None

    if filename.endswith(('.txt', '.csv')):
        try:
            decoded_content = file_content.decode('utf-8')
        except UnicodeDecodeError:
            decoded_content = file_content.decode('latin-1') # Fallback for other encodings

        # Try to parse as CSV first, then fallback to line-by-line
        reader = csv.reader(io.StringIO(decoded_content))
        for row in reader:
            for item in row:
                # A simple regex to detect email-like strings
                email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', item)
                if email_match:
                    emails.append(email_match.group(0))
        if not emails: # If CSV parsing didn't find emails, try line by line
            for line in decoded_content.splitlines():
                email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', line)
                if email_match:
                    emails.append(email_match.group(0))

    elif filename.endswith(('.xlsx', '.xls')):
        if not EXCEL_SUPPORT:
            print("Excel file detected but openpyxl is not installed. Cannot process.")
            raise HTTPException(status_code=400, detail="Excel file processing requires 'openpyxl'. Please install it (`pip install openpyxl`).")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            sheet = workbook.active
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', cell.value)
                        if email_match:
                            emails.append(email_match.group(0))
        except Exception as e:
            print(f"Error processing Excel file: {e}")
            raise HTTPException(status_code=400, detail=f"Error reading Excel file: {e}. Ensure it's a valid Excel format.")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Please upload .txt, .csv, or .xlsx files.")

    # Deduplicate emails and return
    return list(dict.fromkeys(emails))


# UI Endpoint to serve the HTML file
@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    """
    Serves the main HTML file for file uploads.
    """
    return templates.TemplateResponse("upload.html", {"request": request})

# File Upload Endpoint
@app.post("/uploadfile")
async def create_upload_file(file: UploadFile = File(...)):
    """
    Handles uploaded email list files, extracts emails, and processes each one.
    """
    print(f"Received file upload: {file.filename}, Content-Type: {file.content_type}")

    if not (file.filename.endswith(('.txt', '.csv', '.xlsx', '.xls'))):
        raise HTTPException(status_code=400, detail="Invalid file type. Only .txt, .csv, .xlsx, and .xls files are supported.")

    file_contents = await file.read()
    emails_to_process = []
    processed_count = 0
    failed_count = 0
    errors = []

    try:
        emails_to_process = extract_emails_from_file_content(file_contents, file.filename)
        if not emails_to_process:
            raise HTTPException(status_code=400, detail="No valid email addresses found in the uploaded file.")
        
        print(f"Found {len(emails_to_process)} emails to process.")

        for email_address in emails_to_process:
            print(f"Processing email from uploaded file: {email_address}")
            current_email_status = "failed" # Default status
            current_error_message = None
            email_subject_to_log = "N/A" # Default for logging if not generated
            company_domain_to_log = "N/A"
            recipient_email_to_log = TEST_RECEIVER_EMAIL

            try:
                # --- Existing Logic from /process-email endpoint, modified slightly ---
                parsed_info = parse_email(email_address)
                if not parsed_info["domain"] or not parsed_info["first_name"]:
                    print(f"Skipping email '{email_address}': Could not parse name/domain.")
                    current_error_message = f"Could not parse name/domain for {email_address}."
                    failed_count += 1
                    errors.append(current_error_message)
                    db_manager.log_email_processing(
                        original_email=email_address,
                        recipient_email_sent_to=recipient_email_to_log,
                        company_domain=parsed_info["domain"] if parsed_info["domain"] else "N/A",
                        email_subject="Parsing Failed",
                        status="failed",
                        error_message=current_error_message
                    )
                    continue

                recipient_name = parsed_info["first_name"]
                company_domain = parsed_info["domain"]
                company_domain_to_log = company_domain # Update for logging
                
                # actual_recipient_email = TEST_RECEIVER_EMAIL # Already set above as recipient_email_to_log

                company_name = parsed_info["full_name"].split(" ")[-1] if parsed_info["full_name"] else company_domain.split('.')[0].capitalize()

                target_url = get_full_url(company_domain)
                scraped_data = scrape_website_content(target_url)

                scraped_title = None
                scraped_meta_description = None
                scraped_main_content = ""

                if scraped_data["error"]:
                    print(f"Scraping error for {target_url}: {scraped_data['error']}")
                    current_error_message = f"Scraping failed for {company_domain}: {scraped_data['error']}"
                    errors.append(current_error_message)
                else:
                    scraped_title = scraped_data['title']
                    scraped_meta_description = scraped_data['meta_description']
                    scraped_main_content = scraped_data['main_content'] or ""

                llm_input_content = f"Title: {scraped_title or ''}\n" \
                                    f"Description: {scraped_meta_description or ''}\n" \
                                    f"Content: {scraped_main_content or ''}"
                
                if not llm_processor.generator:
                    print("LLM not initialized, skipping content generation.")
                    current_error_message = f"LLM not available for {email_address}."
                    errors.append(current_error_message)
                    failed_count += 1
                    db_manager.log_email_processing(
                        original_email=email_address,
                        recipient_email_sent_to=recipient_email_to_log,
                        company_domain=company_domain_to_log,
                        email_subject="LLM Init Failed",
                        status="failed",
                        error_message=current_error_message
                    )
                    continue

                business_summary = llm_processor.summarize_business(company_name, llm_input_content)
                improvement_tips = llm_processor.generate_improvement_tips(company_name, business_summary, llm_input_content)
                email_body = llm_processor.generate_email_body(recipient_name, company_name, business_summary, improvement_tips)

                email_data = {
                    "recipient_name": recipient_name,
                    "company_name": company_name,
                    "business_summary": business_summary,
                    "improvement_tips": improvement_tips,
                    "email_body": email_body,
                    "sender_name": "Dave Chatpar",
                    "sender_company": "NetWit",
                    "reply_to_email": REPLY_TO_EMAIL
                }
                html_email_content = email_generator.generate_email_html(email_data)

                email_subject = f"Personalized Insights for Your Business: {company_name}"
                email_subject_to_log = email_subject # Update for logging

                if not email_sender:
                    print("Email sender not configured, skipping email sending.")
                    current_error_message = f"Email sender not configured for {email_address}."
                    errors.append(current_error_message)
                    failed_count += 1
                    db_manager.log_email_processing(
                        original_email=email_address,
                        recipient_email_sent_to=recipient_email_to_log,
                        company_domain=company_domain_to_log,
                        email_subject=email_subject_to_log,
                        status="failed",
                        error_message=current_error_message
                    )
                    continue

                email_sent_success = email_sender.send_email(
                    sender_email=SENDER_EMAIL,
                    recipient_email=recipient_email_to_log,
                    reply_to_email=REPLY_TO_EMAIL,
                    subject=email_subject,
                    html_content=html_email_content
                )

                if email_sent_success:
                    processed_count += 1
                    current_email_status = "success"
                else:
                    failed_count += 1
                    current_email_status = "failed"
                    current_error_message = f"Failed to send email to {recipient_email_to_log} for {email_address}. Check SMTP logs."
                    errors.append(current_error_message)

            except Exception as e:
                print(f"Error processing individual email '{email_address}': {e}")
                failed_count += 1
                current_email_status = "failed"
                current_error_message = f"Processing failed for email {email_address}: {e}"
                errors.append(current_error_message)
            finally:
                # Log the outcome for EACH email, even if an exception occurred within the loop
                db_manager.log_email_processing(
                    original_email=email_address,
                    recipient_email_sent_to=recipient_email_to_log,
                    company_domain=company_domain_to_log,
                    email_subject=email_subject_to_log,
                    status=current_email_status,
                    error_message=current_error_message
                )

        # --- End of Existing Logic ---

        message = f"Successfully processed {processed_count} emails. "
        if failed_count > 0:
            message += f"Failed to process {failed_count} emails. See details below."
        elif processed_count == 0 and failed_count == 0 and len(emails_to_process) > 0:
             message = "No emails were processed. Please check your file content and server logs for errors."
        elif len(emails_to_process) == 0:
            message = "The uploaded file did not contain any valid email addresses."


        return {
            "message": message,
            "total_emails_in_file": len(emails_to_process),
            "successfully_processed": processed_count,
            "failed_to_process": failed_count,
            "errors": errors
        }

    except HTTPException as e:
        raise e # Re-raise FastAPI HTTPExceptions
    except Exception as e:
        print(f"Overall file processing error: {e}")
        raise HTTPException(status_code=500, detail=f"An error occurred during file processing: {e}")

# Keep the original /process-email endpoint for single email processing (optional, but good for API consistency)
@app.post("/process-email")
async def process_single_email_endpoint(request: EmailProcessRequest):
    """
    Processes a single email address:
    1. Parses the email to extract name and domain.
    2. Scrapes the associated company website.
    3. Uses an LLM to generate business summary and improvement tips.
    4. Generates a personalized HTML email.
    5. Sends the email via SMTP.
    (This endpoint also logs to the database)
    """
    email_address = request.email
    print(f"Received request to process single email: {email_address}")

    current_email_status = "failed" # Default status for single email
    current_error_message = None
    email_subject_to_log = "N/A"
    company_domain_to_log = "N/A"
    recipient_email_to_log = TEST_RECEIVER_EMAIL # Will be used in finally block


    try:
        # 1. Parse Email
        parsed_info = parse_email(email_address)
        if not parsed_info["domain"] or not parsed_info["first_name"]:
            current_error_message = "Could not parse email address or extract domain/name."
            raise HTTPException(status_code=400, detail=current_error_message)

        recipient_name = parsed_info["first_name"]
        company_domain = parsed_info["domain"]
        company_domain_to_log = company_domain
        
        company_name = parsed_info["full_name"].split(" ")[-1] if parsed_info["full_name"] else company_domain.split('.')[0].capitalize()


        print(f"Parsed Info: Name='{recipient_name}', Domain='{company_domain}', Company='{company_name}'")

        # 2. Web Scraping
        target_url = get_full_url(company_domain)
        print(f"Attempting to scrape URL: {target_url}")
        scraped_data = scrape_website_content(target_url)

        scraped_title = None
        scraped_meta_description = None
        scraped_main_content = ""

        if scraped_data["error"]:
            print(f"Scraping error encountered, proceeding with minimal content: {scraped_data['error']}")
            current_error_message = f"Scraping error: {scraped_data['error']}" # Log scraping errors
        else:
            scraped_title = scraped_data['title']
            scraped_meta_description = scraped_data['meta_description']
            scraped_main_content = scraped_data['main_content'] or ""

        llm_input_content = f"Title: {scraped_title or ''}\n" \
                            f"Description: {scraped_meta_description or ''}\n" \
                            f"Content: {scraped_main_content or ''}"

        print(f"Scraped content length for LLM: {len(llm_input_content)} characters.")

        # 3. LLM Integration
        if not llm_processor.generator:
            current_error_message = "LLM model not initialized. Check server logs."
            raise HTTPException(status_code=500, detail=current_error_message)

        print("Generating business summary with LLM...")
        business_summary = llm_processor.summarize_business(company_name, llm_input_content)
        print(f"Business Summary: {business_summary}")

        print("Generating improvement tips with LLM...")
        improvement_tips = llm_processor.generate_improvement_tips(company_name, business_summary, llm_input_content)
        print(f"Improvement Tips: {improvement_tips}")

        print("Generating email body with LLM...")
        email_body = llm_processor.generate_email_body(recipient_name, company_name, business_summary, improvement_tips)
        print(f"Email Body (first 200 chars): {email_body[:200]}...")


        # 4. Generate HTML Email
        email_data = {
            "recipient_name": recipient_name,
            "company_name": company_name,
            "business_summary": business_summary,
            "improvement_tips": improvement_tips,
            "email_body": email_body,
            "sender_name": "Sushanth K.S.",
            "sender_company": "NothingRightNow",
            "reply_to_email": REPLY_TO_EMAIL
        }
        html_email_content = email_generator.generate_email_html(email_data)
        print("HTML email content generated.")

        email_subject = f"Personalized Insights for Your Business: {company_name}"
        email_subject_to_log = email_subject # Update for logging

        # 5. Send Email
        if not email_sender:
            current_error_message = "Email sender not configured. Check .env variables."
            raise HTTPException(status_code=500, detail=current_error_message)

        email_sent_success = email_sender.send_email(
            sender_email=SENDER_EMAIL,
            recipient_email=recipient_email_to_log,
            reply_to_email=REPLY_TO_EMAIL,
            subject=email_subject,
            html_content=html_email_content
        )

        if email_sent_success:
            current_email_status = "success"
            return {
                "status": "success",
                "message": "Email processed and sent successfully!",
                "parsed_info": parsed_info,
                "scraped_data_summary": {
                    "title": scraped_title,
                    "meta_description": scraped_meta_description,
                    "content_length": len(scraped_main_content)
                },
                "llm_generated_content": {
                    "business_summary": business_summary,
                    "improvement_tips": improvement_tips,
                    "email_body_preview": email_body[:150] + "..."
                },
                "email_sent_to": recipient_email_to_log
            }
        else:
            current_error_message = "Failed to send email. Check server logs for details."
            raise HTTPException(status_code=500, detail=current_error_message)

    except HTTPException as http_exc:
        current_error_message = http_exc.detail
        raise http_exc # Re-raise FastAPI HTTPExceptions
    except Exception as e:
        print(f"Overall processing error for single email '{email_address}': {e}")
        current_error_message = f"An unexpected error occurred: {e}"
        raise HTTPException(status_code=500, detail=current_error_message)
    finally:
        # Log the outcome for the single email, even if an exception occurred
        db_manager.log_email_processing(
            original_email=email_address,
            recipient_email_sent_to=recipient_email_to_log,
            company_domain=company_domain_to_log,
            email_subject=email_subject_to_log,
            status=current_email_status,
            error_message=current_error_message
        )

# NEW: Endpoint to retrieve all processed email logs
@app.get("/logs")
async def get_processed_email_logs():
    """
    Retrieves all logs of processed emails from the SQLite database.
    """
    print("Retrieving all processed email logs from database.")
    try:
        logs = db_manager.get_all_logs()
        return {"logs": logs}
    except Exception as e:
        print(f"Error retrieving logs: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve logs: {e}")
